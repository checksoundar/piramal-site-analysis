#!/usr/bin/env python3
"""Build WesternSydney_URLs.xlsx — Western Sydney University URL discovery."""
import json
import os
from collections import Counter
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
CATALOG = os.path.join(BASE, "url-discovery", "www-westernsydney-edu-au")

HEADER_FILL = PatternFill("solid", fgColor="A00030")   # WSU deep red
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, size=14, color="A00030")

d = json.load(open(os.path.join(CATALOG, "urls-all.json")))["analysis-urls-all"]
urls = sorted(u["url"] for u in d["urls"])
docs = sorted(u["url"] for u in d.get("documents", []))


def section(u):
    segs = [s for s in urlparse(u).path.split("/") if s]
    return segs[0] if segs else "(home)"


def subsection(u):
    segs = [s for s in urlparse(u).path.split("/") if s]
    if not segs:
        return "(home)"
    return "/" + "/".join(segs[:2])


wb = Workbook()

# ---- Summary ----
ws = wb.active
ws.title = "Summary"
ws.append(["Western Sydney University — URL Discovery"])
ws["A1"].font = TITLE
ws.append(["https://www.westernsydney.edu.au/"])
ws.append([])
rows = [
    ("Total pages (URLs)", d["totalUrls"]),
    ("Documents", d["totalDocuments"]),
    ("Discovery method", d["method"]),
    ("Sitemap", d.get("sitemapURL", "")),
    ("Confidence", d.get("confidence", "")),
    ("Captured", (d.get("captured", "") or "")[:19].replace("T", " ")),
    ("Distinct top-level sections", len({section(u) for u in urls})),
    ("Notes", d.get("limitations", "")),
]
for k, v in rows:
    ws.append([k, v])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 95

# ---- Sections breakdown ----
ws2 = wb.create_sheet("Sections")
ws2.append(["Top-level section", "Page count"])
for c in range(1, 3):
    ws2.cell(row=1, column=c).fill = HEADER_FILL
    ws2.cell(row=1, column=c).font = HEADER_FONT
    ws2.cell(row=1, column=c).alignment = Alignment(horizontal="center")
for sec, n in Counter(section(u) for u in urls).most_common():
    ws2.append([sec, n])
ws2.column_dimensions["A"].width = 45
ws2.column_dimensions["B"].width = 14
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:B{ws2.max_row}"

# ---- All URLs ----
ws3 = wb.create_sheet("All URLs")
ws3.append([f"All pages ({d['totalUrls']})  —  method: {d['method']}, sitemap: {d.get('sitemapURL','')}"])
ws3["A1"].font = Font(bold=True, size=12, color="A00030")
ws3.append([])
ws3.append(["#", "URL", "Section", "Sub-section"])
for c in range(1, 5):
    ws3.cell(row=3, column=c).fill = HEADER_FILL
    ws3.cell(row=3, column=c).font = HEADER_FONT
    ws3.cell(row=3, column=c).alignment = Alignment(horizontal="center")
for i, u in enumerate(urls, 1):
    ws3.append([i, u, section(u), subsection(u)])
    c = ws3.cell(row=ws3.max_row, column=2)
    c.hyperlink = u
    c.font = Font(color="0563C1", underline="single")
ws3.column_dimensions["A"].width = 7
ws3.column_dimensions["B"].width = 95
ws3.column_dimensions["C"].width = 26
ws3.column_dimensions["D"].width = 40
ws3.freeze_panes = "A4"
ws3.auto_filter.ref = f"A3:D{ws3.max_row}"

# ---- Documents (if any) ----
if docs:
    ws4 = wb.create_sheet("Documents")
    ws4.append(["#", "Document URL"])
    for c in range(1, 3):
        ws4.cell(row=1, column=c).fill = HEADER_FILL
        ws4.cell(row=1, column=c).font = HEADER_FONT
    for i, u in enumerate(docs, 1):
        ws4.append([i, u])
        ws4.cell(row=ws4.max_row, column=2).hyperlink = u
    ws4.column_dimensions["A"].width = 7
    ws4.column_dimensions["B"].width = 100
    ws4.freeze_panes = "A2"

out = os.path.join(BASE, "WesternSydney_URLs.xlsx")
wb.save(out)
print("Saved:", out)
print("Sheets:", [s.title for s in wb.worksheets])
print("Pages:", d["totalUrls"], "| Sections:", len({section(u) for u in urls}))
print("Size: %.1f KB" % (os.path.getsize(out) / 1024))
