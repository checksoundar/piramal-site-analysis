#!/usr/bin/env python3
"""Build an Excel workbook with the urls-all list for each site on a separate sheet."""
import json
import os
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

# (folder, sheet name, site URL) — order = discovery order, largest-first grouping
SITES = [
    ("id-mamypoko",   "id.mamypoko.com",   "https://id.mamypoko.com"),
    ("jp-unicharmpet", "jp.unicharmpet.com", "https://jp.unicharmpet.com"),
    ("unicharm-cojp", "www.unicharm.co.jp", "https://www.unicharm.co.jp"),
    ("jp-moony",      "jp.moony.com",      "https://jp.moony.com"),
    ("www-sofy-jp",   "www.sofy.jp",       "https://www.sofy.jp"),
    ("jp-lifree",     "jp.lifree.com",     "https://jp.lifree.com"),
]
# folder key aliases (sofy-jp folder is "sofy-jp")
FOLDER_ALIAS = {"www-sofy-jp": "sofy-jp"}

HEADER_FILL = PatternFill("solid", fgColor="006CB5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUM_HDR_FILL = PatternFill("solid", fgColor="1F4E79")


def section_of(url, site_url):
    """Second path segment after the locale prefix, as a rough content-section label."""
    path = urlparse(url).path
    segs = [s for s in path.split("/") if s]
    if not segs:
        return "(home)"
    locales = {"ja", "en", "id", "zh", "ko", "th", "vi", "pt", "es", "ar"}
    if segs[0] in locales:
        if len(segs) == 1:
            return "(locale root)"
        return segs[0] + "/" + segs[1].replace(".html", "")
    return segs[0].replace(".html", "")


def load(folder):
    path = os.path.join(BASE, folder, "urls-all.json")
    data = json.load(open(path))["analysis-urls-all"]
    return data


wb = Workbook()

# ---- Summary sheet (first) ----
ws = wb.active
ws.title = "Summary"
ws.append(["Unicharm — URL Discovery Summary"])
ws["A1"].font = Font(bold=True, size=14, color="006CB5")
ws.append([])
hdr = ["#", "Sheet / Host", "Site URL", "Total URLs", "Documents", "Method", "Sitemap", "Confidence", "Captured"]
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    cell = ws.cell(row=3, column=c)
    cell.fill = SUM_HDR_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")

summ = []
grand = 0
for i, (folder, sheet, url) in enumerate(SITES, 1):
    fk = FOLDER_ALIAS.get(folder, folder)
    d = load(fk)
    grand += d["totalUrls"]
    summ.append((folder, sheet, url, d))
    ws.append([
        i, sheet, url, d["totalUrls"], d["totalDocuments"], d["method"],
        d.get("sitemapURL", ""), d.get("confidence", ""), d.get("captured", "")[:19].replace("T", " "),
    ])
ws.append([])
total_row = ws.max_row + 1
ws.append(["", "TOTAL", "", grand])
ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

widths = [4, 22, 34, 11, 11, 10, 22, 11, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A4"

# ---- Per-site sheets ----
for folder, sheet, url in SITES:
    fk = FOLDER_ALIAS.get(folder, folder)
    d = load(fk)
    urls = [u["url"] for u in d["urls"]]
    urls.sort()
    ws = wb.create_sheet(title=sheet[:31])  # Excel sheet-name max 31 chars
    # title/meta rows
    ws.append([f"{sheet}  —  {d['totalUrls']} URLs  (method: {d['method']}, sitemap: {d.get('sitemapURL','')})"])
    ws["A1"].font = Font(bold=True, size=12, color="006CB5")
    ws.append([])
    ws.append(["#", "URL", "Section"])
    for c in range(1, 4):
        cell = ws.cell(row=3, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for idx, u in enumerate(urls, 1):
        ws.append([idx, u, section_of(u, url)])
        ws.cell(row=ws.max_row, column=2).hyperlink = u
        ws.cell(row=ws.max_row, column=2).font = Font(color="0563C1", underline="single")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 26
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:C{ws.max_row}"

out = os.path.join(BASE, "Unicharm_Site_URLs.xlsx")
wb.save(out)
print("Saved:", out)
print("Sheets:", [s.title for s in wb.worksheets])
print("Grand total URLs:", grand)
print("Size: %.1f KB" % (os.path.getsize(out) / 1024))
