#!/usr/bin/env python3
"""Build Unicharm_Batch2_URLs.xlsx — one sheet per batch-2 site (urls-all list)."""
import json
import os
import glob
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

HEADER_FILL = PatternFill("solid", fgColor="006CB5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUM_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
LOCALES = {"ja", "en", "id", "zh", "ko", "th", "vi", "pt", "es", "ar", "zh-tw", "my"}


def section_of(url):
    segs = [s for s in urlparse(url).path.split("/") if s]
    if not segs:
        return "(home)"
    if segs[0] in LOCALES:
        if len(segs) == 1:
            return "(locale root)"
        return segs[0] + "/" + segs[1].replace(".html", "")
    return segs[0].replace(".html", "")


def sheet_name(host):
    """Excel sheet names max 31 chars and cannot contain : \\ / ? * [ ]."""
    n = host
    for ch in ':\\/?*[]':
        n = n.replace(ch, "-")
    return n[:31]


# Discover all site folders that have urls-all.json
sites = []
for folder in sorted(glob.glob(os.path.join(BASE, "*", "urls-all.json"))):
    d = json.load(open(folder))["analysis-urls-all"]
    host = urlparse(d["urls"][0]["url"]).netloc if d["urls"] else os.path.basename(os.path.dirname(folder))
    sites.append((os.path.dirname(folder), host, d))

# Sort by URL count descending
sites.sort(key=lambda x: -x[2]["totalUrls"])

wb = Workbook()

# ---- Summary sheet ----
ws = wb.active
ws.title = "Summary"
ws.append(["Unicharm — Batch 2 URL Discovery (54 sites)"])
ws["A1"].font = Font(bold=True, size=14, color="006CB5")
ws.append([])
hdr = ["#", "Sheet / Host", "Total URLs", "Documents", "Method", "Sitemap", "Confidence", "Captured"]
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    cell = ws.cell(row=3, column=c)
    cell.fill = SUM_HDR_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")

grand = 0
used_names = {}
for i, (folder, host, d) in enumerate(sites, 1):
    grand += d["totalUrls"]
    ws.append([
        i, host, d["totalUrls"], d["totalDocuments"], d["method"],
        d.get("sitemapURL") or "-", d.get("confidence", ""),
        (d.get("captured", "") or "")[:19].replace("T", " "),
    ])
ws.append([])
ws.append(["", "TOTAL", grand])
ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
for i, w in enumerate([4, 24, 11, 11, 16, 30, 11, 20], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A4"

# ---- Per-site sheets ----
for folder, host, d in sites:
    urls = sorted(u["url"] for u in d["urls"])
    name = sheet_name(host)
    # ensure uniqueness
    if name in used_names:
        used_names[name] += 1
        name = f"{name[:28]}_{used_names[name]}"
    else:
        used_names[name] = 1
    ws = wb.create_sheet(title=name)
    ws.append([f"{host}  —  {d['totalUrls']} URLs  (method: {d['method']}, sitemap: {d.get('sitemapURL') or '-'})"])
    ws["A1"].font = Font(bold=True, size=12, color="006CB5")
    ws.append([])
    ws.append(["#", "URL", "Section"])
    for c in range(1, 4):
        cell = ws.cell(row=3, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for idx, u in enumerate(urls, 1):
        ws.append([idx, u, section_of(u)])
        c = ws.cell(row=ws.max_row, column=2)
        c.hyperlink = u
        c.font = Font(color="0563C1", underline="single")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 26
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:C{ws.max_row}"

out = os.path.join(BASE, "Unicharm_Batch2_URLs.xlsx")
wb.save(out)
print("Saved:", out)
print("Sheets:", len(wb.worksheets), "(1 summary +", len(sites), "sites)")
print("Grand total URLs:", grand)
print("Size: %.1f KB" % (os.path.getsize(out) / 1024))
